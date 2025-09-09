#!/usr/bin/env python3
"""
Script para analizar la estructura de la plantilla_despensa_n8n.xlsx
"""

import pandas as pd
import openpyxl
import os
import sys

def analyze_excel_template(filename):
    """Analizar la estructura del archivo Excel"""
    print(f"📊 ANÁLISIS DE PLANTILLA: {filename}")
    print("=" * 60)
    
    if not os.path.exists(filename):
        print(f"❌ Archivo no encontrado: {filename}")
        return None
    
    try:
        # Cargar el archivo Excel
        workbook = openpyxl.load_workbook(filename)
        
        print(f"📋 Hojas disponibles: {workbook.sheetnames}")
        
        structure_info = {}
        
        # Analizar cada hoja
        for sheet_name in workbook.sheetnames:
            print(f"\n🔍 Analizando hoja: '{sheet_name}'")
            print("-" * 40)
            
            # Usar pandas para leer los datos
            try:
                df = pd.read_excel(filename, sheet_name=sheet_name)
                
                print(f"📏 Dimensiones: {df.shape[0]} filas x {df.shape[1]} columnas")
                print(f"📑 Columnas: {list(df.columns)}")
                
                # Mostrar primeras filas si hay datos
                if not df.empty:
                    print(f"\n📝 Primeras 3 filas:")
                    print(df.head(3).to_string())
                    
                    # Analizar tipos de datos
                    print(f"\n🔢 Tipos de datos:")
                    for col in df.columns:
                        non_null_count = df[col].notna().sum()
                        dtype = df[col].dtype
                        print(f"   {col}: {dtype} ({non_null_count} valores no nulos)")
                
                else:
                    print("📝 Hoja vacía o solo con encabezados")
                
                # Guardar información de la estructura
                structure_info[sheet_name] = {
                    'columns': list(df.columns),
                    'shape': df.shape,
                    'dtypes': {col: str(df[col].dtype) for col in df.columns}
                }
                
            except Exception as e:
                print(f"⚠️ Error leyendo la hoja '{sheet_name}': {e}")
                
                # Intentar leer directamente con openpyxl
                try:
                    worksheet = workbook[sheet_name]
                    print(f"📏 Dimensiones (openpyxl): {worksheet.max_row} filas x {worksheet.max_column} columnas")
                    
                    # Leer primera fila (posibles encabezados)
                    headers = []
                    for col in range(1, worksheet.max_column + 1):
                        cell_value = worksheet.cell(row=1, column=col).value
                        headers.append(cell_value)
                    
                    print(f"📑 Posibles encabezados: {headers}")
                    
                    structure_info[sheet_name] = {
                        'headers': headers,
                        'max_row': worksheet.max_row,
                        'max_column': worksheet.max_column
                    }
                    
                except Exception as e2:
                    print(f"❌ Error con openpyxl: {e2}")
        
        return structure_info
        
    except Exception as e:
        print(f"❌ Error general analizando el archivo: {e}")
        return None

def generate_sheets_structure_from_template(structure_info):
    """Generar código de estructura para Google Sheets basado en la plantilla"""
    if not structure_info:
        return
    
    print(f"\n🔧 CÓDIGO SUGERIDO PARA GOOGLE SHEETS")
    print("=" * 60)
    
    for sheet_name, info in structure_info.items():
        print(f"\n# Estructura para hoja: {sheet_name}")
        
        if 'columns' in info:
            columns = info['columns']
            print(f"headers_{sheet_name.lower()} = {columns}")
        elif 'headers' in info:
            headers = [h for h in info['headers'] if h is not None]
            print(f"headers_{sheet_name.lower()} = {headers}")
        
        print(f"# Dimensiones originales: {info.get('shape', 'N/A')}")

def suggest_detection_mapping(structure_info):
    """Sugerir mapeo de detecciones basado en la plantilla"""
    print(f"\n💡 SUGERENCIAS DE MAPEO PARA DETECCIONES")
    print("=" * 60)
    
    for sheet_name, info in structure_info.items():
        columns = info.get('columns', info.get('headers', []))
        if not columns:
            continue
            
        print(f"\nHoja: {sheet_name}")
        print(f"Columnas disponibles: {columns}")
        
        # Sugerir mapeo inteligente
        suggested_mapping = {}
        
        for col in columns:
            col_lower = str(col).lower() if col else ""
            
            if any(word in col_lower for word in ['fecha', 'time', 'date']):
                suggested_mapping[col] = 'timestamp'
            elif any(word in col_lower for word in ['producto', 'item', 'objeto', 'alimento']):
                suggested_mapping[col] = 'class_name'
            elif any(word in col_lower for word in ['categoria', 'tipo', 'category']):
                suggested_mapping[col] = 'category'
            elif any(word in col_lower for word in ['cantidad', 'count', 'numero']):
                suggested_mapping[col] = 'quantity'
            elif any(word in col_lower for word in ['estado', 'status']):
                suggested_mapping[col] = 'status'
            elif any(word in col_lower for word in ['confianza', 'confidence']):
                suggested_mapping[col] = 'confidence'
        
        if suggested_mapping:
            print("Mapeo sugerido:")
            for excel_col, detection_field in suggested_mapping.items():
                print(f"   '{excel_col}' -> {detection_field}")

if __name__ == "__main__":
    filename = "plantilla_despensa_n8n.xlsx"
    
    # Analizar la plantilla
    structure = analyze_excel_template(filename)
    
    if structure:
        # Generar código sugerido
        generate_sheets_structure_from_template(structure)
        
        # Sugerir mapeo de detecciones
        suggest_detection_mapping(structure)
        
        print(f"\n🎯 PRÓXIMOS PASOS:")
        print("1. Revisa la estructura analizada arriba")
        print("2. Modifica google_sheets_integration.py con la estructura correcta")
        print("3. Ajusta el mapeo de detecciones según tus necesidades")
        print("4. Prueba la integración con test_sheets_integration.py")
    
    else:
        print("\n❌ No se pudo analizar la plantilla")
        print("💡 Verifica que el archivo exista y sea accesible")
